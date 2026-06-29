from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path


def iter_xlsx_rows(path: Path, *, sheet_name: str | None = None) -> Iterator[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(values_only=True):
            yield ["" if cell is None else str(cell) for cell in row]
    finally:
        wb.close()


def read_xlsx_header(path: Path, **kwargs) -> list[str]:
    for row in iter_xlsx_rows(path, **kwargs):
        return [cell.strip() for cell in row]
    return []
